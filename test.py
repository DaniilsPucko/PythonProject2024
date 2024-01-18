import pandas as pd
import selenium
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

def read_txt_file(english):
    word_list = []

    try:
        with open(english, 'r') as file:
            content = file.read()
            words = content.split(',')
            word_list.extend(word.strip() for word in words)

        return word_list

    except FileNotFoundError:
        print(f"File not found: {english}")
        return None

def translate_to_language(words, target_language):
    translated_dict = {}
    driver = webdriver.Chrome()  
    driver.get(f"https://www.bing.com/translator?from=en&to={target_language}&setlang=be")
    
    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'tta_input_ta'))
        )
    finally:
        for word in words:
            input_field = driver.find_element(By.ID, 'tta_input_ta')
            input_field.clear()
            input_field.send_keys(word)
            time.sleep(2)

            try:
                find = driver.find_element(By.ID, 'tta_output_ta')
                translation = find.get_attribute("value")

                print(f"{word} (English) -> {translation} ({target_language.capitalize()})")
                translated_dict[word] = translation

            except Exception as e:
                print(f"Translation failed for {word}: {e}")

            time.sleep(2)

    driver.quit()
    return translated_dict

file_path = 'english.txt'
words_to_translate = read_txt_file(file_path)

if words_to_translate is not None:
    translated_words_latvian = translate_to_language(words_to_translate, 'lv')
    translated_words_french = translate_to_language(words_to_translate, 'fr')
    
    if translated_words_latvian and translated_words_french:
        excel_file_path = 'translated_words.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value='English Word')
        sheet.cell(row=1, column=2, value='Latvian Translation')
        sheet.cell(row=1, column=3, value='French Translation')

        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 50


        for i, (word, translation_latvian) in enumerate(translated_words_latvian.items(), start=2):
            translation_french = translated_words_french.get(word, '')
            sheet.cell(row=i, column=1, value=word)
            sheet.cell(row=i, column=2, value=translation_latvian)
            sheet.cell(row=i, column=3, value=translation_french)

        workbook.save(excel_file_path)
        print(f"Translated words saved to {excel_file_path}")
    else:
        print("No words found or translation unsuccessful.")

else:
    print("No words found.")